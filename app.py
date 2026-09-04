import streamlit as st
import pandas as pd
import time
import os
import unicodedata
from io import BytesIO
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from pypdf import PdfReader
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. Estrutura de Extração
class ItemLicitacao(BaseModel):
    numero_item: str = Field(description="Número ou identificador do item/lote no edital (ex: '01', 'Lote 1 - Item 2')")
    descricao: str = Field(description="Descrição completa do medicamento no edital (princípio ativo, dosagem, forma)")
    principio_ativo_identificado: str = Field(description="Substância / Denominação genérica do princípio ativo identificado")
    unidade: str = Field(description="Unidade de medida/fornecimento (ex: AMP, FA, COMP, FR)")
    quantidade: float = Field(description="Quantidade demandada expressa em número float")
    valor_referencia_unitario: float = Field(description="Valor unitário máximo ou de referência do edital")

class ListaItens(BaseModel):
    itens: list[ItemLicitacao]

# 2. Configurações da Página
st.set_page_config(page_title="Radar Farma - Licitações", layout="wide", page_icon="💊")
st.title("🎯 Analisador de Editais & Mapa de Preços")
st.caption("Varredura inteligente de editais cruzada diretamente com o banco de dados oficial de portfólio.")

# Chave de API
if "GEMINI_API_KEY" in st.secrets and st.secrets["GEMINI_API_KEY"]:
    api_key = st.secrets["GEMINI_API_KEY"]
else:
    api_key = st.sidebar.text_input("Chave Gemini API", type="password")

# 3. Base de Dados e Normalização
ARQUIVO_PORTFOLIO = "portfolio_laboratorios.xlsx"

def normalizar_texto(texto):
    if not texto or pd.isna(texto):
        return ""
    nfkd = unicodedata.normalize('NFKD', str(texto))
    texto_sem_acento = "".join([c for c in nfkd if not unicodedata.combining(c)])
    return texto_sem_acento.upper().strip()

MAPEAMENTO_LABS = {
    "Blau": ["BLAU"],
    "Eurofarma": ["EUROFARMA"],
    "Baxter": ["BAXTER"],
    "Biocon": ["BIOCON"],
    "Accord": ["ACCORD"],
    "Halex Istar": ["HALEX", "ISTAR"],
    "United Medical": ["UNITED MEDICAL", "UNITED"],
    "GSK": ["GSK", "GLAXO", "GLAXOSMITHKLINE"],
    "Aspen": ["ASPEN"],
    "Sanofi": ["SANOFI"],
    "Pint Pharma": ["PINT", "PINT PHARMA"]
}

@st.cache_data
def carregar_base_portfolio(caminho):
    if not os.path.exists(caminho):
        return None
    try:
        df = pd.read_excel(caminho)
        df.columns = [str(c).strip().upper() for c in df.columns]
        for col in ["SUBSTÂNCIA", "LABORATÓRIO", "PRODUTO", "APRESENTAÇÃO"]:
            if col in df.columns:
                df[col] = df[col].astype(str).str.strip()
        return df
    except Exception:
        return None

df_portfolio = carregar_base_portfolio(ARQUIVO_PORTFOLIO)

# 4. Barra Lateral
st.sidebar.header("⚙️ Configurações de Busca")

segmento = st.sidebar.selectbox(
    "Segmento do Edital:",
    options=["Medicamentos", "Material Elétrico / Engenharia", "Personalizado / Geral"]
)

if segmento == "Medicamentos":
    st.sidebar.subheader("💊 Portfólio de Laboratórios")
    labs_base = [
        "Blau", "Eurofarma", "Baxter", "Biocon", "Accord",
        "Halex Istar", "United Medical", "GSK", "Aspen", "Sanofi", "Pint Pharma"
    ]
    
    labs_selecionados = st.sidebar.multiselect(
        "Laboratórios ativos para cotação:",
        options=labs_base,
        default=labs_base
    )
    
    if df_portfolio is not None:
        st.sidebar.success(f"Base Conectada: {len(df_portfolio)} produtos cadastrados.")
    else:
        st.sidebar.warning("Arquivo 'portfolio_laboratorios.xlsx' não encontrado na raiz.")

    filtro_exibicao = st.sidebar.radio(
        "Visualização dos Resultados:",
        ["Apenas Itens com Match nos Laboratórios", "Todos os Medicamentos do Edital"],
        index=0
    )

    st.sidebar.info(
        "**Hierarquia Comercial Fixada:**\n"
        "- 🥇 **Sanofi** (Referência) tem prioridade máxima.\n"
        "- 🥈 **Blau** tem prioridade sobre **Eurofarma**.\n"
        "- 🚫 **Sanofi:** exclui linha Medley (sem genéricos)."
    )

elif segmento == "Material Elétrico / Engenharia":
    portfolio_texto = st.sidebar.text_area(
        "Itens de interesse:",
        value="Material elétrico, cabeamento estruturado, iluminação LED, quadros de distribuição"
    )
else:
    portfolio_texto = st.sidebar.text_area("Itens de interesse:", value="")

arquivo_pdf = st.file_uploader("Arraste o PDF do Edital ou Termo de Referência aqui", type=["pdf"])

def extrair_texto_pdf(arquivo_carregado):
    leitor = PdfReader(arquivo_carregado)
    texto_total = []
    for i, pagina in enumerate(leitor.pages):
        conteudo = pagina.extract_text()
        if conteudo:
            texto_total.append(f"--- PÁGINA {i+1} ---\n{conteudo}")
    return "\n\n".join(texto_total)

def enriquecer_com_portfolio(df_extraido, df_port, labs_escolhidos):
    if df_port is None:
        df_extraido["Laboratório Sugerido"] = "Sem base carregada"
        df_extraido["Produto / Marca Ref."] = "-"
        return df_extraido

    base = df_port.copy()
    base["SUBSTANCIA_NORM"] = base["SUBSTÂNCIA"].apply(normalizar_texto)
    base["LABORATORIO_NORM"] = base["LABORATÓRIO"].apply(normalizar_texto)
    base["PRODUTO_NORM"] = base["PRODUTO"].apply(normalizar_texto)

    termos_busca_labs = []
    for lab in labs_escolhidos:
        termos_busca_labs.extend(MAPEAMENTO_LABS.get(lab, [normalizar_texto(lab)]))

    # Regra Sanofi: Bloqueio de Medley / Genéricos
    mascara_medley = base["PRODUTO_NORM"].str.contains("MEDLEY") | \
                     base["LABORATORIO_NORM"].str.contains("MEDLEY")
    base_valida = base[~mascara_medley]

    labs_atribuidos = []
    produtos_atribuidos = []

    for _, row in df_extraido.iterrows():
        substancia_edital = normalizar_texto(row["Princípio Ativo"])
        primeira_palavra = substancia_edital.split()[0] if substancia_edital else ""

        matches = base_valida[
            base_valida["SUBSTANCIA_NORM"].str.contains(substancia_edital, regex=False, na=False) |
            base_valida["SUBSTANCIA_NORM"].apply(lambda s: primeira_palavra in s if len(primeira_palavra) > 4 else False)
        ]

        matches = matches[matches["LABORATORIO_NORM"].apply(
            lambda lab_nome: any(termo in lab_nome for termo in termos_busca_labs)
        )]

        if not matches.empty:
            labs_encontrados = matches["LABORATORIO_NORM"].unique().tolist()
            
            # Hierarquia: Sanofi > Blau > Eurofarma
            tem_sanofi = any("SANOFI" in l for l in labs_encontrados)
            tem_blau = any("BLAU" in l for l in labs_encontrados)
            tem_euro = any("EUROFARMA" in l for l in labs_encontrados)

            if tem_sanofi:
                lab_final_filtro = [l for l in labs_encontrados if "SANOFI" in l][0]
            elif tem_blau:
                lab_final_filtro = [l for l in labs_encontrados if "BLAU" in l][0]
            elif tem_euro:
                lab_final_filtro = [l for l in labs_encontrados if "EUROFARMA" in l][0]
            else:
                lab_final_filtro = labs_encontrados[0]

            linha_escolhida = matches[matches["LABORATORIO_NORM"] == lab_final_filtro].iloc[0]
            labs_atribuidos.append(linha_escolhida["LABORATÓRIO"])
            produtos_atribuidos.append(linha_escolhida["PRODUTO"])
        else:
            labs_atribuidos.append("Não mapeado / Verificar")
            produtos_atribuidos.append("-")

    df_extraido["Laboratório Sugerido"] = labs_atribuidos
    df_extraido["Produto / Marca Ref."] = produtos_atribuidos
    return df_extraido

# Gerador Excel Executivo com Fórmulas e Estilos
def gerar_excel_estilizado(df_dados):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_dados.to_excel(writer, index=False, sheet_name="Mapa_de_Precos")
        ws = writer.sheets["Mapa_de_Precos"]
        ws.views.sheetView[0].showGridLines = True

        fonte_cabecalho = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        fill_cabecalho = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        
        fonte_corpo = Font(name="Calibri", size=10)
        fill_editavel = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        alinhamento_centro = Alignment(horizontal="center", vertical="center")
        alinhamento_esquerda = Alignment(horizontal="left", vertical="center")
        alinhamento_direita = Alignment(horizontal="right", vertical="center")
        
        borda_fina = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9")
        )

        num_linhas = len(df_dados)
        num_cols = len(df_dados.columns)

        # 1. Cabeçalho
        for col_num in range(1, num_cols + 1):
            celula = ws.cell(row=1, column=col_num)
            celula.font = fonte_cabecalho
            celula.fill = fill_cabecalho
            celula.alignment = alinhamento_centro
            ws.row_dimensions[1].height = 28

        # 2. Dados e Fórmulas
        for row_idx in range(2, num_linhas + 2):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = fonte_corpo
                cell.border = borda_fina

                if col_idx in [1, 4]:  # Item e Unidade
                    cell.alignment = alinhamento_centro
                elif col_idx in [2, 3, 6, 7]:  # Descrição, Princípio, Lab, Produto
                    cell.alignment = alinhamento_esquerda
                elif col_idx == 5:  # Quantidade
                    cell.alignment = alinhamento_direita
                    cell.number_format = "#,##0"
                elif col_idx == 6:  # Valor Ref. Unitário
                    cell.alignment = alinhamento_direita
                    cell.number_format = "R$ #,##0.00"
                elif col_idx == 8:  # Custo Aquisição (Editável)
                    cell.alignment = alinhamento_direita
                    cell.number_format = "R$ #,##0.00"
                    cell.fill = fill_editavel
                elif col_idx == 9:  # Margem Alvo (%)
                    cell.alignment = alinhamento_direita
                    cell.number_format = "0.0%"
                    cell.value = 0.15  # 15%
                elif col_idx == 10:  # Preço Proposta Unitário (Fórmula Dinâmica)
                    cell.alignment = alinhamento_direita
                    cell.number_format = "R$ #,##0.00"
                    cell.value = f"=H{row_idx}*(1+I{row_idx})"

        # 3. Autoajuste de Colunas
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        ws.column_dimensions['B'].width = 42
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['F'].width = 28
        ws.column_dimensions['G'].width = 22

    return output.getvalue()

# 5. Processamento
if arquivo_pdf and api_key:
    if st.button("🚀 Processar Edital", type="primary"):
        with st.spinner("Extraindo itens e gerando Mapa de Preços executivo..."):
            try:
                texto_edital = extrair_texto_pdf(arquivo_pdf)

                if not texto_edital.strip():
                    st.error("O PDF parece ser uma imagem digitalizada sem camada de texto pesquisável.")
                    st.stop()

                guia_substancias = ""
                if segmento == "Medicamentos" and df_portfolio is not None:
                    subs_unicas = df_portfolio["SUBSTÂNCIA"].dropna().unique()[:250].tolist()
                    guia_substancias = f"Lista de referência de substâncias prioritárias:\n[{', '.join(subs_unicas)}]"

                client = genai.Client(api_key=api_key)

                if segmento == "Medicamentos":
                    prompt = f"""
                    Você é um analista sênior de licitações farmacêuticas e compras hospitalares.
                    Analise o texto do edital e extraia TODOS os itens de medicamentos licitados.

                    Laboratórios alvo: [{", ".join(labs_selecionados)}]
                    {guia_substancias}

                    DIRETRIZES MANDATÓRIAS:
                    1. Identifique e extraia todos os itens de medicamentos com descrição, dosagem e forma farmacêutica.
                    2. No campo 'principio_ativo_identificado', extraia a denominação genérica (ex: 'Propofol', 'Enoxaparina Sódica', 'Meropenem').
                    3. Converta quantidades e valores de referência para números decimais (float).
                    4. Retorne no formato JSON estruturado.

                    TEXTO DO EDITAL:
                    \"\"\"
                    {texto_edital}
                    \"\"\"
                    """
                else:
                    prompt = f"""
                    Você é um analista de licitações. Extraia os itens correspondentes a: "{portfolio_texto}".
                    Estruture no JSON solicitado com quantidades e valores unitários numéricos float.

                    TEXTO DO EDITAL:
                    \"\"\"
                    {texto_edital}
                    \"\"\"
                    """

                modelos_tentativa = ["gemini-3.6-flash", "gemini-2.5-flash", "gemini-1.5-flash-latest"]
                resposta = None
                ultimo_erro = None

                for nome_modelo in modelos_tentativa:
                    try:
                        resposta = client.models.generate_content(
                            model=nome_modelo,
                            contents=prompt,
                            config=types.GenerateContentConfig(
                                response_mime_type="application/json",
                                response_schema=ListaItens,
                            )
                        )
                        if resposta:
                            break
                    except Exception as err:
                        ultimo_erro = err
                        time.sleep(2)

                if not resposta:
                    raise ultimo_erro

                dados = ListaItens.model_validate_json(resposta.text)

                if not dados.itens:
                    st.warning("Nenhum item foi identificado no edital.")
                else:
                    df = pd.DataFrame([item.model_dump() for item in dados.itens])
                    df.columns = [
                        "Item", 
                        "Descrição Completa Edital", 
                        "Princípio Ativo", 
                        "Unidade", 
                        "Qtd", 
                        "Valor Ref. Unit. (R$)"
                    ]

                    if segmento == "Medicamentos":
                        df = enriquecer_com_portfolio(df, df_portfolio, labs_selecionados)
                        if filtro_exibicao == "Apenas Itens com Match nos Laboratórios":
                            df = df[df["Laboratório Sugerido"] != "Não mapeado / Verificar"]

                    df["Custo Aquisição (R$)"] = 0.0
                    df["Margem Alvo (%)"] = 0.15
                    df["Preço Proposta Unit. (R$)"] = 0.0

                    st.success(f"Foram identificados e mapeados {len(df)} itens no edital!")
                    st.dataframe(df, use_container_width=True)

                    excel_bytes = gerar_excel_estilizado(df)

                    st.download_button(
                        label="📥 Baixar Mapa de Preços Profissional (.xlsx)",
                        data=excel_bytes,
                        file_name=f"Mapa_Precos_{arquivo_pdf.name.replace('.pdf', '')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except Exception as e:
                st.error(f"Erro no processamento: {str(e)}")

elif not api_key:
    st.info("Insira a sua chave de API na barra lateral ou configure nos secrets para iniciar a análise.")